import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# COLOUR PALETTE
# ─────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"
C_MID_BLUE    = "2E75B6"
C_LIGHT_BLUE  = "D6E4F0"
C_SKY         = "EBF5FB"
C_GREEN       = "1E8449"
C_AMBER       = "D4AC0D"
C_RED_DARK    = "C0392B"
C_YELLOW_LITE = "FEF9E7"
C_WHITE       = "FFFFFF"
C_ORANGE      = "D35400"
C_GREY_LITE   = "F2F3F4"
C_DARK_GREY   = "2C3E50"

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def wrap_center(horizontal="center", vertical="center", wrap=True):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="A0A0A0")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="1F3864")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_row, max_row, min_col, max_col, border_fn=thin_border):
    b = border_fn()
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(r, c).border = b

def merge_title(ws, cell_range, text, bg, font_obj, align=None):
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    top_left.value = text
    top_left.fill = hex_fill(bg)
    top_left.font = font_obj
    top_left.alignment = align or wrap_center()

def section_header(ws, row, col_start, col_end, text,
                   bg=C_MID_BLUE, font_color=C_WHITE):
    cell_range = f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}"
    ws.merge_cells(cell_range)
    cell = ws.cell(row, col_start)
    cell.value = text
    cell.fill = hex_fill(bg)
    cell.font = Font(name="Calibri", size=11, bold=True, color=font_color)
    cell.alignment = wrap_center("left")
    apply_border(ws, row, row, col_start, col_end)

def col_header(ws, row, col, text, bg=C_DARK_BLUE, font_color=C_WHITE, size=10):
    cell = ws.cell(row, col)
    cell.value = text
    cell.fill = hex_fill(bg)
    cell.font = Font(name="Calibri", size=size, bold=True, color=font_color)
    cell.alignment = wrap_center()
    cell.border = thin_border()

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER / PROJECT INFO
# ═══════════════════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "Cover and Project Info"
ws_cover.sheet_view.showGridLines = False

for col, w in [(1,4),(2,28),(3,35),(4,20),(5,20),(6,4)]:
    ws_cover.column_dimensions[get_column_letter(col)].width = w
for r in range(1, 80):
    ws_cover.row_dimensions[r].height = 18

ws_cover.row_dimensions[2].height = 50
ws_cover.row_dimensions[3].height = 30
ws_cover.row_dimensions[4].height = 22

merge_title(ws_cover,"B2:E2",
    "METEOROLOGICAL MAST - PROCUREMENT QUALITY CHECKLIST",
    C_DARK_BLUE, Font(name="Calibri",size=18,bold=True,color=C_WHITE))

merge_title(ws_cover,"B3:E3",
    "Lattice Structure | Wind Resource Assessment | India",
    C_MID_BLUE, Font(name="Calibri",size=12,bold=False,color=C_WHITE))

merge_title(ws_cover,"B4:E4",
    "Template Version 2.0  |  Prepared by: Wind Projects Engineering Team",
    C_DARK_GREY, Font(name="Calibri",size=10,bold=False,color=C_LIGHT_BLUE))

r = 6
section_header(ws_cover, r, 2, 5, "  PROJECT INFORMATION", C_DARK_GREY)

fields = [
    ("Project Name",          "Rajasthan Wind Resource Assessment - Phase 1"),
    ("Project Location",      "Jaisalmer District, Rajasthan, India"),
    ("Client / Developer",    "Green Horizon Energy Pvt. Ltd."),
    ("EPC / Consultant",      "WindTech Engineering Solutions Pvt. Ltd."),
    ("Document Reference",    "GHEL-JSLM-MM-PRQ-001"),
    ("Mast Height",           "120 m (Lattice, Free-Standing)"),
    ("No. of Masts",          "2 Nos."),
    ("Site Coordinates",      "Lat: 27.0833 N  |  Long: 70.9000 E"),
    ("Terrain Category",      "Category II - Open Flat Terrain (IS 875 Part 3)"),
    ("Wind Zone",             "Wind Zone V - Basic Wind Speed 50 m/s (IS 875)"),
    ("Procurement Stage",     "RFQ / Technical Bid Evaluation"),
    ("Expected Delivery",     "16 weeks from Purchase Order"),
    ("Installation Start",    "Q3 2025"),
    ("Monitoring Duration",   "Minimum 12 months (target 24 months)"),
    ("Prepared By",           "Arun Sharma - Lead Wind Resource Engineer"),
    ("Reviewed By",           "Priya Nair - Senior Structural Engineer"),
    ("Date of Issue",         "2025-06-01"),
    ("Document Status",       "DRAFT FOR REVIEW"),
]

for i,(label,value) in enumerate(fields):
    row = r+1+i
    ws_cover.row_dimensions[row].height = 20
    lc = ws_cover.cell(row,2)
    lc.value = label
    lc.fill = hex_fill(C_LIGHT_BLUE if i%2==0 else C_SKY)
    lc.font = Font(name="Calibri",size=10,bold=True,color=C_DARK_BLUE)
    lc.alignment = wrap_center("left")
    lc.border = thin_border()
    vc = ws_cover.cell(row,3)
    vc.value = value
    vc.fill = hex_fill(C_LIGHT_BLUE if i%2==0 else C_SKY)
    vc.font = body_font(10)
    vc.alignment = wrap_center("left")
    vc.border = thin_border()
    ws_cover.merge_cells(f"C{row}:E{row}")

idx_row = r + len(fields) + 3
section_header(ws_cover, idx_row, 2, 5, "  DOCUMENT INDEX", C_DARK_GREY)
sheets_index = [
    ("Sheet 1","Cover and Project Info",       "Project metadata and document control"),
    ("Sheet 2","Vendor Qualification",         "Pre-qualification criteria and scoring"),
    ("Sheet 3","Mast Structure Tech Spec",     "Structural design, material, loads"),
    ("Sheet 4","Foundation and Civil",         "Foundation design and civil requirements"),
    ("Sheet 5","Fabrication and QA-QC",        "Shop fabrication checklist and inspection"),
    ("Sheet 6","Surface Protection",           "Galvanizing, coating, corrosion protection"),
    ("Sheet 7","Erection and Safety",          "Field erection, safety requirements"),
    ("Sheet 8","Instrumentation",              "Sensors, data logger, cabling"),
    ("Sheet 9","Documentation",               "Drawings, certifications, reports"),
    ("Sheet 10","Commercial Terms",            "Warranty, delivery, payment, LD"),
    ("Sheet 11","Bid Evaluation Matrix",       "Weighted scoring for vendor comparison"),
]
col_header(ws_cover, idx_row+1, 2, "Sheet")
col_header(ws_cover, idx_row+1, 3, "Title")
col_header(ws_cover, idx_row+1, 4, "Content Summary")
col_header(ws_cover, idx_row+1, 5, "Status")

for i,(sh,title,desc) in enumerate(sheets_index):
    rr = idx_row+2+i
    ws_cover.row_dimensions[rr].height = 18
    bg = C_SKY if i%2 else C_WHITE
    for col,val in [(2,sh),(3,title),(4,desc),(5,"Included")]:
        cc = ws_cover.cell(rr,col)
        cc.value = val
        cc.fill = hex_fill(bg)
        cc.font = body_font(10)
        cc.alignment = wrap_center("left")
        cc.border = thin_border()

ab_row = idx_row + len(sheets_index) + 4
section_header(ws_cover, ab_row, 2, 5, "  ABBREVIATIONS", C_DARK_GREY)
abbrevs = [
    ("IS","Indian Standard"),("IEC","International Electrotechnical Commission"),
    ("MET","Meteorological"),("WRA","Wind Resource Assessment"),
    ("NDA","Non-Disclosure Agreement"),("QA","Quality Assurance"),
    ("QC","Quality Control"),("MTC","Mill Test Certificate"),
    ("FAT","Factory Acceptance Test"),("SAT","Site Acceptance Test"),
    ("RFQ","Request for Quotation"),("PO","Purchase Order"),
    ("LD","Liquidated Damages"),("HAT","Hub Height Anemometry Tower"),
    ("BOM","Bill of Materials"),("PPE","Personal Protective Equipment"),
]
col_header(ws_cover, ab_row+1, 2, "Abbreviation")
col_header(ws_cover, ab_row+1, 3, "Full Form")

for i,(a,b) in enumerate(abbrevs):
    rr = ab_row+2+i
    ws_cover.row_dimensions[rr].height = 17
    bg = C_SKY if i%2 else C_WHITE
    for col,val in [(2,a),(3,b)]:
        cc = ws_cover.cell(rr,col)
        cc.value = val
        cc.fill = hex_fill(bg)
        cc.font = body_font(10, bold=(col==2))
        cc.alignment = wrap_center("left")
        cc.border = thin_border()
    ws_cover.merge_cells(f"C{rr}:E{rr}")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — VENDOR QUALIFICATION
# ═══════════════════════════════════════════════════════════════════════════
ws_vq = wb.create_sheet("Vendor Qualification")
ws_vq.sheet_view.showGridLines = False

for col,w in {1:3,2:6,3:40,4:30,5:25,6:18,7:18,8:3}.items():
    ws_vq.column_dimensions[get_column_letter(col)].width = w

merge_title(ws_vq,"B1:G1","VENDOR PRE-QUALIFICATION CHECKLIST - MET MAST (LATTICE STRUCTURE)",
            C_DARK_BLUE, Font(name="Calibri",size=14,bold=True,color=C_WHITE))
merge_title(ws_vq,"B2:G2",
    "Project: Rajasthan WRA Ph-1 | Client: Green Horizon Energy Pvt. Ltd. | Doc: GHEL-JSLM-MM-PRQ-001",
    C_MID_BLUE, Font(name="Calibri",size=10,bold=False,color=C_WHITE))

ws_vq.row_dimensions[3].height = 40
merge_title(ws_vq,"B3:G3",
    "PURPOSE: This sheet captures the minimum qualification criteria a vendor must meet before "
    "receiving the RFQ. All Mandatory items must be satisfied. Score Preferred items for ranking.",
    C_YELLOW_LITE, Font(name="Calibri",size=9,italic=True,color=C_DARK_GREY))

r=4
for col,hdr in {2:"#",3:"Qualification Criterion",4:"Requirement / Benchmark",
                5:"Vendor Response / Evidence",6:"Mandatory / Preferred",
                7:"Compliance (Y/N/Partial)"}.items():
    col_header(ws_vq,r,col,hdr)
ws_vq.row_dimensions[r].height=30

vq_sections = {
    "A. COMPANY PROFILE AND EXPERIENCE":[
        ("A1","Years in business - lattice tower / telecom / met mast fabrication",
         "Minimum 10 years in structural steel fabrication",
         "Copy of Certificate of Incorporation + Company profile","Mandatory"),
        ("A2","Number of met masts / lattice towers fabricated and supplied",
         "Minimum 15 met masts (>=80 m height) in last 5 years",
         "Client reference list with tower heights, year, contact","Mandatory"),
        ("A3","Experience with projects in high wind zone / desert terrain",
         "At least 3 projects in Wind Zone IV or V (IS 875 Part 3)",
         "Project data sheets, site photographs","Preferred"),
        ("A4","Experience supplying to renewable energy developers",
         "Wind energy developers or IPPs as past clients",
         "Reference list including developer names","Preferred"),
        ("A5","In-house structural engineering team",
         "At least 2 qualified structural engineers (B.E./B.Tech Civil/Structural)",
         "Team CVs, educational qualifications","Mandatory"),
    ],
    "B. CERTIFICATIONS AND QUALITY SYSTEM":[
        ("B1","ISO 9001:2015 - Quality Management System",
         "Valid, scope covering structural steel fabrication",
         "Copy of current ISO certificate","Mandatory"),
        ("B2","ISO 3834 - Welding quality requirements",
         "ISO 3834-2 or 3834-3 preferred",
         "Certificate copy","Preferred"),
        ("B3","IS 2062 / IS 808 material procurement compliance",
         "Steel sourced from SAIL/TATA/JSPL or equivalent",
         "Mill test certificates (MTC) for sample projects","Mandatory"),
        ("B4","Third-party inspection empanelment",
         "Must accept Bureau Veritas / TUV / DNV inspection",
         "Written confirmation in bid","Mandatory"),
        ("B5","CE Marking or IEC compliance for tower structure",
         "IEC 61400-1 structural loads reference preferred",
         "Declaration or design report reference","Preferred"),
        ("B6","Compliance with Factories Act and labour laws",
         "Valid factory licence, ESI/PF registration",
         "Copies of valid licences","Mandatory"),
    ],
    "C. MANUFACTURING FACILITY":[
        ("C1","In-house fabrication workshop area",
         "Minimum 5,000 sq.m covered fabrication area",
         "Factory layout drawings, photographs","Mandatory"),
        ("C2","CNC / plasma / laser cutting machines",
         "Minimum 1 CNC plasma/laser cutting machine",
         "Equipment list with make and capacity","Preferred"),
        ("C3","Drilling and punching machines for angle / plate",
         "Radial drill / CNC drill, capacity >= 32 mm dia",
         "Equipment list","Mandatory"),
        ("C4","Overhead crane capacity",
         "Minimum 10-tonne EOT crane in assembly bay",
         "Equipment register","Mandatory"),
        ("C5","Hot-dip galvanizing facility (in-house or tie-up)",
         "BS EN ISO 1461 / IS 4759 compliant; tie-up acceptable",
         "GI facility cert or MOU with sub-contractor","Mandatory"),
        ("C6","Welding equipment and certified welders",
         "MIG/MAG/SMAW; welders certified per IS 7310",
         "Welder qualification records (WQR)","Mandatory"),
        ("C7","NDT testing capability",
         "UT / MPI / DPT in-house or third-party empanelled",
         "NDT agency credentials","Mandatory"),
        ("C8","Pre-assembly / trial erection area",
         "Adequate yard for pre-assembly of full mast sections",
         "Site photographs","Preferred"),
        ("C9","Material traceability system",
         "Heat-number based traceability from mill to dispatch",
         "Procedure document","Mandatory"),
    ],
    "D. DESIGN CAPABILITY":[
        ("D1","In-house structural analysis software",
         "STAAD.Pro / SAP2000 / Tekla or equivalent",
         "Software licences / sample analysis report","Mandatory"),
        ("D2","Signed and stamped structural design report",
         "Licensed Structural Engineer (Member of IStructE / IE(I))",
         "Sample report from past project","Mandatory"),
        ("D3","Foundation design capability or sub-consultant",
         "Geotechnical input integration; spread / pile foundation design",
         "Past foundation design reports","Preferred"),
        ("D4","Wind load calculation per IS 875 Part 3 / IEC 61400-1",
         "Design basis document referencing applicable codes",
         "Sample design basis document","Mandatory"),
        ("D5","Fatigue analysis capability",
         "S-N curve based fatigue life assessment for lattice members",
         "Sample or reference","Preferred"),
        ("D6","Deflection / tilt analysis",
         "Maximum allowable tip deflection criteria stated",
         "Analysis excerpt from past project","Preferred"),
    ],
    "E. FINANCIAL AND COMMERCIAL STANDING":[
        ("E1","Annual turnover",
         "Minimum INR 10 Crore (avg. last 3 financial years)",
         "Audited P&L for last 3 years","Mandatory"),
        ("E2","Net worth positive",
         "Positive net worth in latest audited balance sheet",
         "Audited balance sheet","Mandatory"),
        ("E3","Bank solvency / credit line",
         "Bank solvency letter equivalent to contract value",
         "Bank letter","Mandatory"),
        ("E4","No litigation / blacklisting",
         "Self-declaration not blacklisted by any GoI body",
         "Affidavit","Mandatory"),
        ("E5","Ability to provide Performance Bank Guarantee",
         "10% of PO value; validity PO duration + 6 months",
         "Confirmation in bid","Mandatory"),
    ],
    "F. HSE - HEALTH, SAFETY AND ENVIRONMENT":[
        ("F1","Dedicated HSE officer in factory",
         "Full-time qualified safety officer",
         "Appointment letter, qualification","Mandatory"),
        ("F2","Lost Time Injury frequency (LTIF) rate",
         "LTIF < 1.0 per million man-hours (last 3 years)",
         "HSE statistics report","Preferred"),
        ("F3","HSE plan for field erection",
         "Site-specific HSE plan to be submitted before erection",
         "Template HSE plan","Mandatory"),
        ("F4","PPE compliance",
         "Mandatory PPE for all workers during fabrication and erection",
         "PPE policy document","Mandatory"),
    ],
}

row = 5
alt = False
for section, items in vq_sections.items():
    section_header(ws_vq, row, 2, 7, f"  {section}")
    row += 1
    for item in items:
        ws_vq.row_dimensions[row].height = 40
        bg = C_SKY if alt else C_WHITE
        for col,val in [(2,item[0]),(3,item[1]),(4,item[2]),(5,""),(6,item[4]),(7,"")]:
            cc = ws_vq.cell(row,col)
            cc.value = val
            cc.fill = hex_fill(bg)
            cc.font = body_font(10)
            cc.alignment = wrap_center("left")
            cc.border = thin_border()
            if col==6:
                if val=="Mandatory":
                    cc.font = Font(name="Calibri",size=10,bold=True,color=C_RED_DARK)
                else:
                    cc.font = Font(name="Calibri",size=10,bold=True,color=C_GREEN)
        alt = not alt
        row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — MAST STRUCTURE TECHNICAL SPECIFICATION
# ═══════════════════════════════════════════════════════════════════════════
ws_st = wb.create_sheet("Mast Structure Tech Spec")
ws_st.sheet_view.showGridLines = False

for col,w in {1:3,2:5,3:38,4:35,5:28,6:22,7:20,8:18,9:3}.items():
    ws_st.column_dimensions[get_column_letter(col)].width = w

merge_title(ws_st,"B1:H1",
    "METEOROLOGICAL MAST - STRUCTURAL TECHNICAL SPECIFICATION CHECKLIST",
    C_DARK_BLUE, Font(name="Calibri",size=14,bold=True,color=C_WHITE))
merge_title(ws_st,"B2:H2",
    "Project: Rajasthan WRA Ph-1  |  Mast Height: 120 m  |  Type: Self-Supporting Lattice  |  Wind Zone: V",
    C_MID_BLUE, Font(name="Calibri",size=10,color=C_WHITE))

r=3
for col,hdr in {2:"Ref",3:"Parameter / Question to Vendor",
                4:"Client Requirement / Benchmark",
                5:"Vendor Proposed Value / Response",
                6:"Applicable Code / Standard",
                7:"Importance",8:"Compliance (Y/N/Partial)"}.items():
    col_header(ws_st,r,col,hdr)
    ws_st.row_dimensions[r].height=32

struct_data = {
    "1. GENERAL MAST CONFIGURATION":[
        ("1.1","What is the proposed mast type?",
         "Self-supporting lattice steel tower; no guy wires; free-standing",
         "","IS 875 Part 3, IEC 61400-1","Critical"),
        ("1.2","What is the total hub/sensor height?",
         "120 m above finished ground level (FGL); confirm with survey benchmark",
         "","IS 875 Part 3","Critical"),
        ("1.3","What is the base width (footprint) of the lattice at foundation level?",
         "Minimum 10-12% of total height (>=12 m base width for 120 m); confirm calculation",
         "","Structural analysis output","Critical"),
        ("1.4","What are the proposed section heights and number of sections?",
         "Sections <=12 m for transportability; state number and flange splice locations",
         "","Erection methodology","High"),
        ("1.5","What is the design wind speed used for structural design?",
         "V_b = 50 m/s (Zone V); K1=1.08 (50yr return); K2, K3 as per terrain",
         "","IS 875 Part 3:2015","Critical"),
        ("1.6","Has gust factor / dynamic analysis been performed?",
         "Gust factor method or dynamic analysis; resonance check required",
         "","IS 875 Part 3 Cl. 8","High"),
        ("1.7","What is the maximum allowable tip deflection?",
         "<= H/150 under design wind load (<= 800 mm for 120 m)",
         "","IS 800:2007 Cl. 5","High"),
        ("1.8","What is the design life of the mast structure?",
         "Minimum 25 years; fatigue loading considered for repeating wind cycles",
         "","IEC 61400-1 Ed.4","Critical"),
        ("1.9","What is the maximum mast tilt / plumb tolerance after installation?",
         "<= 1:500 of height (<= 240 mm lateral lean) measured at top",
         "","IEC 61400-12-1","Critical"),
        ("1.10","Is a pre-erection trial assembly planned at factory?",
         "Yes - full mast trial assembly mandatory before dispatch; photos + survey report",
         "","QA/QC requirement","High"),
    ],
    "2. STRUCTURAL STEEL - MATERIAL SPECIFICATION":[
        ("2.1","What steel grade is proposed for main leg members (chord)?",
         "IS 2062 E350 BR (FU min 490 MPa, FY min 350 MPa); Charpy impact >=27J @ 0 deg C",
         "","IS 2062:2011","Critical"),
        ("2.2","What steel grade for bracing / diagonal members?",
         "IS 2062 E250 A minimum; E350 preferred for higher-load bracings",
         "","IS 2062:2011","High"),
        ("2.3","What section profiles are used for legs, bracings, and platforms?",
         "Legs: rolled steel equal/unequal angles or SHS/RHS; Bracings: angles or rods; state BOM",
         "","IS 808","High"),
        ("2.4","How is material traceability maintained from mill to final product?",
         "Heat number stamping / colour coding; Mill Test Certificates (MTC) for every heat",
         "","ISO 10474 / IS 2062","Critical"),
        ("2.5","What is the carbon equivalent (CE) of the proposed steel?",
         "CE <= 0.43% for weldability (IS 2062 E350); confirm weld procedure qualification",
         "","IS 2062","High"),
        ("2.6","Is the steel sourced from primary (integrated) steel plants?",
         "SAIL / TATA Steel / JSW / JSPL or equivalent; secondary/re-rolled NOT acceptable",
         "","Client requirement","Critical"),
        ("2.7","What is the plate/section thickness range for leg and gusset plates?",
         "State min and max thickness; ensure thickness >= 8 mm for main legs at base",
         "","Structural drawing","High"),
        ("2.8","Are all bolts, nuts, and washers hot-dip galvanized high-tensile?",
         "Grade 8.8 bolts HDG per IS 1367 Part 13; property class 8 nuts; hardened washers",
         "","IS 1367, IS 1367-3","Critical"),
        ("2.9","Are flanged bolted joints or welded splices used for section joints?",
         "Flanged bolted joints preferred for on-site assembly; HSFG bolts at flanges",
         "","IS 800 Cl.10.4","High"),
        ("2.10","What is the proposed bolt diameter and pitch at main splice flanges?",
         "State bolt size (M24/M27/M30), grade, pitch, edge distance; preload torque values",
         "","IS 800","High"),
    ],
    "3. STRUCTURAL DESIGN LOADS":[
        ("3.1","What are the dead loads considered?",
         "Provide load summary: self-weight, instrument load, ice/snow if applicable, cable load",
         "","IS 875 Part 1","Critical"),
        ("3.2","What wind loads are applied at each section height?",
         "Show wind pressure distribution (kN/m2) at 10 m intervals; solidity ratio of lattice faces",
         "","IS 875 Part 3 Cl. 6","Critical"),
        ("3.3","Are instrument drag loads included in structural design?",
         "Anemometer, wind vane, boom arm drag; minimum 0.5 kN horizontal load per boom level",
         "","IEC 61400-12-1","High"),
        ("3.4","What is the design seismic zone and seismic coefficient used?",
         "Rajasthan - Zone III; seismic load per IS 1893 Part 1",
         "","IS 1893:2016","High"),
        ("3.5","Is a combined wind + seismic load combination checked?",
         "Per IS 800 load combinations; governing load case identified",
         "","IS 800:2007 Cl. 5","High"),
        ("3.6","What factor of safety (FOS) is applied to structural members?",
         "FOS >= 1.5 on yield; member utilisation ratio <= 0.90 under ultimate load",
         "","IS 800","Critical"),
        ("3.7","Is fatigue life assessed for high-cycle wind loading?",
         "IEC 61400-1 fatigue load classes; S-N curves for welded joints; design life >= 25 yr",
         "","IEC 61400-1 Ed.4","High"),
        ("3.8","Are guy wire attachment loads included (if applicable)?",
         "N/A for self-supporting; if temporary guys used during erection, loads must be stated",
         "","Erection plan","Medium"),
    ],
    "4. SECTION GEOMETRY AND PLATFORM DETAILS":[
        ("4.1","What are the cross-arm / boom arm specifications?",
         "Boom length: min 2.0 m from mast face; state boom tube O.D., wall thickness, material",
         "","IEC 61400-12-1","Critical"),
        ("4.2","How many instrument platforms / service platforms are provided?",
         "Min 2 intermediate access platforms (at ~40 m and ~80 m); top instrument platform",
         "","IEC 61400-12-1","High"),
        ("4.3","What are the platform dimensions and load rating?",
         "Min 0.6 m width; load rating >= 2.5 kN/m2 live load; anti-slip grating",
         "","IS 800","High"),
        ("4.4","What fall protection / safety cage / rest platforms are provided?",
         "Full height safety ladder with safety cage >= 2.5 m or fall-arrest rail (EN 353-1)",
         "","EN 353-1, IS 3696","Critical"),
        ("4.5","What is the ladder rung spacing and side rail dimensions?",
         "Rung spacing: 300 mm; side rail width: min 300 mm; galvanized steel throughout",
         "","IS 3696","High"),
        ("4.6","How are cable and conduit runs managed on the mast?",
         "Dedicated cable tray / conduit brackets at max 2 m spacing; earthing cable route",
         "","Client requirement","High"),
        ("4.7","Are aviation warning lights and mounting provisions required?",
         "Yes - aviation obstruction light provision per DGCA / AAI requirements at top and mid height",
         "","DGCA CAR Section 4","Critical"),
        ("4.8","What is the top plate configuration for instrument mounting?",
         "Instrument mounting ring / top plate with pre-drilled holes for anemometer and vane booms",
         "","IEC 61400-12-1","High"),
    ],
    "5. EARTHING AND LIGHTNING PROTECTION":[
        ("5.1","Is a dedicated earthing and lightning protection system provided with the mast?",
         "Yes - fully designed earthing per IS 3043; lightning protection per IS/IEC 62305",
         "","IS 3043, IEC 62305","Critical"),
        ("5.2","What is the proposed earthing electrode system?",
         "Minimum 4 nos. copper-bonded earth rods (3 m x 17 mm); ring electrode around foundation",
         "","IS 3043","Critical"),
        ("5.3","What is the target earth resistance value?",
         "<= 5 Ohm (measured after installation); <= 10 Ohm in rocky terrain with chemical treatment",
         "","IS 3043 / IEC 62305","Critical"),
        ("5.4","Are surge protection devices (SPDs) supplied for instruments?",
         "SPDs at data logger cabinet and at each signal/power cable entry - Class II SPD minimum",
         "","IEC 61643-11","Critical"),
        ("5.5","Is the earthing system compatible with data logger cabinet bonding?",
         "All metallic parts bonded; equipotential bonding bar in data logger enclosure",
         "","IEC 62305","High"),
        ("5.6","What type of lightning finial / air terminal is proposed?",
         "Franklin rod type; min 1.5 m above highest instrument; Cu or Al alloy",
         "","IEC 62305-3","High"),
    ],
    "6. PAINTING AND MARKING":[
        ("6.1","What aviation marking colour scheme is proposed?",
         "Red and white alternate bands (600 mm) per DGCA CAR; top section red",
         "","DGCA CAR Section 4","Critical"),
        ("6.2","What is the mast identification marking / nameplate?",
         "SS nameplate welded at base: project name, mast ID, height, year of fabrication, vendor name",
         "","Client requirement","High"),
        ("6.3","Are height interval markers provided on the mast?",
         "Yes - height markers every 10 m for instrument level identification",
         "","Client requirement","Medium"),
    ],
}

row = 4
alt = False
for section, items in struct_data.items():
    section_header(ws_st, row, 2, 8, f"  {section}")
    row += 1
    for item in items:
        ws_st.row_dimensions[row].height = 50
        bg = C_SKY if alt else C_WHITE
        importance_color = {
            "Critical": C_RED_DARK,
            "High":     C_ORANGE,
            "Medium":   C_AMBER,
            "Low":      C_GREEN,
        }.get(item[5], "000000")
        for col, val in [(2,item[0]),(3,item[1]),(4,item[2]),(5,""),(6,item[3]),(7,item[5]),(8,"")]:
            cc = ws_st.cell(row, col)
            cc.value = val
            cc.fill = hex_fill(bg)
            cc.font = body_font(10)
            if col==7:
                cc.font = Font(name="Calibri",size=10,bold=True,color=importance_color)
            cc.alignment = wrap_center("left")
            cc.border = thin_border()
        alt = not alt
        row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 — FOUNDATION AND CIVIL
# ═══════════════════════════════════════════════════════════════════════════
ws_fnd = wb.create_sheet("Foundation and Civil")
ws_fnd.sheet_view.showGridLines = False

for col,w in {1:3,2:5,3:38,4:35,5:28,6:22,7:20,8:18,9:3}.items():
    ws_fnd.column_dimensions[get_column_letter(col)].width = w

merge_title(ws_fnd,"B1:H1","FOUNDATION AND CIVIL WORKS - SPECIFICATION CHECKLIST",
            C_DARK_BLUE, Font(name="Calibri",size=14,bold=True,color=C_WHITE))
merge_title(ws_fnd,"B2:H2",
    "Project: Rajasthan WRA Ph-1 | Soil: Silty Sand / Sandy Soil | Foundation: Spread Footing",
    C_MID_BLUE, Font(name="Calibri",size=10,color=C_WHITE))

r=3
for col,hdr in {2:"Ref",3:"Parameter / Question",4:"Requirement / Benchmark",
                5:"Vendor Response",6:"Standard",7:"Importance",8:"Compliance"}.items():
    col_header(ws_fnd,r,col,hdr)
ws_fnd.row_dimensions[r].height=30

foundation_data = {
    "1. GEOTECHNICAL INVESTIGATION":[
        ("1.1","Has a site-specific geotechnical investigation report been provided?",
         "BH/DCPT/SPT to min 1.5x foundation depth or 6 m; by licensed geo firm",
         "","IS 1892, IS 2131","Critical"),
        ("1.2","What is the allowable bearing capacity (ABC) at foundation depth?",
         "ABC >= 100 kN/m2 at proposed founding level; SBC from plate load test preferred",
         "","IS 6403","Critical"),
        ("1.3","What is the groundwater table depth?",
         "Min 1.5 m below foundation bottom; if shallow, waterproofing and buoyancy check needed",
         "","IS 1892","High"),
        ("1.4","Is soil liquefaction assessment performed (seismic zone III)?",
         "SPT-N >= 15 or liquefaction assessment as per IS 1893 Annex F",
         "","IS 1893","High"),
        ("1.5","What is the soil corrosivity classification?",
         "pH, sulphate, chloride content of soil; if aggressive, use SRPC cement and coated rebars",
         "","IS 456 Annex B","High"),
    ],
    "2. FOUNDATION DESIGN":[
        ("2.1","What foundation type is proposed?",
         "Spread (isolated pad) footing for each leg; raft or pile if soil is poor; justify with analysis",
         "","IS 1904","Critical"),
        ("2.2","What is the design bearing pressure under combined loads?",
         "Factored bearing pressure <= 0.9 x ABC; uplift check under wind/seismic",
         "","IS 1904, IS 456","Critical"),
        ("2.3","What concrete grade is specified for foundation?",
         "M30 minimum for footings; M25 for PCC blinding layer; RCC design as per IS 456",
         "","IS 456","Critical"),
        ("2.4","What is the foundation reinforcement specification?",
         "Fe 500D TMT bars (IS 1786); clear cover >= 75 mm (soil face); 50 mm (top face)",
         "","IS 1786, IS 456","Critical"),
        ("2.5","How are anchor bolts / base plates designed and anchored?",
         "HSFG anchor bolts M36 minimum; embedded depth >= 20x bolt dia; with anchor plate and hook",
         "","IS 456, IS 800","Critical"),
        ("2.6","Is foundation design report stamped by licensed structural engineer?",
         "Yes - design calculations signed by licensed engineer (IE(I) membership) + drawings stamped",
         "","Client requirement","Critical"),
        ("2.7","What is the foundation excavation plan and backfill specification?",
         "Compacted backfill in 150 mm layers; Proctor density >= 95%; no organic material",
         "","IS 2720","High"),
        ("2.8","Is a concrete cube testing plan included?",
         "Min 3 cubes per pour; 7-day and 28-day tests; report to be submitted to client",
         "","IS 516","Critical"),
        ("2.9","Is a foundation settlement monitoring plan included?",
         "Survey benchmarks; settlement reading at 0, 3, 6, 12 months after erection",
         "","IS 1904","High"),
    ],
    "3. CIVIL ANCILLARY WORKS":[
        ("3.1","What is the access road specification to mast location?",
         "All-weather gravel road; min 4 m width; capable of carrying 10T loaded truck",
         "","Client requirement","High"),
        ("3.2","What is the equipment laydown / erection pad specification?",
         "Min 20 m x 20 m compacted area near mast; capable of supporting crane outrigger loads",
         "","Crane erection plan","High"),
        ("3.3","Is a perimeter security fence provided around the mast base?",
         "Min 1.8 m chain link fence with barbed wire top; lockable gate; warning signage in local language",
         "","Client HSE requirement","Critical"),
        ("3.4","What is the data logger enclosure foundation / mounting specification?",
         "RCC plinth (0.3 m above FGL); anti-vibration pads; bolted to plinth",
         "","Client requirement","High"),
        ("3.5","Are cable trenches / conduits provided from mast to data logger?",
         "PVC conduit >= 50 mm ID buried >= 0.6 m; drawstring provided; marked with warning tape",
         "","Client requirement","High"),
        ("3.6","Is site levelling / grading included in scope?",
         "3 m radius around each mast leg graded and compacted; drainage channel if needed",
         "","Client requirement","Medium"),
    ],
}

row = 4
alt = False
for section, items in foundation_data.items():
    section_header(ws_fnd, row, 2, 8, f"  {section}")
    row += 1
    for item in items:
        ws_fnd.row_dimensions[row].height = 45
        bg = C_SKY if alt else C_WHITE
        importance_color = {
            "Critical": C_RED_DARK,"High": C_ORANGE,"Medium": C_AMBER,"Low": C_GREEN
        }.get(item[5],"000000")
        for col, val in [(2,item[0]),(3,item[1]),(4,item[2]),(5,""),(6,item[3]),(7,item[5]),(8,"")]:
            cc = ws_fnd.cell(row, col)
            cc.value = val
            cc.fill = hex_fill(bg)
            cc.font = body_font(10)
            if col==7:
                cc.font = Font(name="Calibri",size=10,bold=True,color=importance_color)
            cc.alignment = wrap_center("left")
            cc.border = thin_border()
        alt = not alt
        row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5 — FABRICATION AND QA/QC
# ═══════════════════════════════════════════════════════════════════════════
ws_qc = wb.create_sheet("Fabrication and QA-QC")
ws_qc.sheet_view.showGridLines = False

for col,w in {1:3,2:5,3:38,4:35,5:28,6:22,7:20,8:18,9:3}.items():
    ws_qc.column_dimensions[get_column_letter(col)].width = w

merge_title(ws_qc,"B1:H1","FABRICATION QUALITY ASSURANCE AND QUALITY CONTROL (QA/QC) CHECKLIST",
            C_DARK_BLUE,Font(name="Calibri",size=14,bold=True,color=C_WHITE))
merge_title(ws_qc,"B2:H2",
    "Project: Rajasthan WRA Ph-1 | Third-Party Inspection: Bureau Veritas India | Stage: Shop Fabrication",
    C_MID_BLUE,Font(name="Calibri",size=10,color=C_WHITE))

r=3
for col,hdr in {2:"Ref",3:"Inspection / QC Item",4:"Acceptance Criterion",
                5:"Inspection Record / Document",6:"Inspection Stage",
                7:"Hold (H) / Witness (W) / Review (R)",8:"Compliance"}.items():
    col_header(ws_qc,r,col,hdr)
ws_qc.row_dimensions[r].height=35

qc_data = {
    "1. INCOMING MATERIAL INSPECTION":[
        ("1.1","Verification of steel mill test certificates (MTC)",
         "Chemical composition and mechanical properties comply with IS 2062; heat number tallies",
         "MTC file (per heat number)","On receipt of material","H"),
        ("1.2","Dimensional check of incoming sections (angle, plate, rod)",
         "Thickness/width within IS 1852 tolerance; cross-check with approved BOM",
         "Material receiving report","On receipt","W"),
        ("1.3","Visual inspection for laminations, cracks, surface defects",
         "Zero laminations; no cracks; minor mill scale acceptable; heavy rust rejected",
         "Visual inspection record","On receipt","W"),
        ("1.4","Verification of bolt, nut, washer certificates",
         "Grade 8.8 property class; IS 1367 compliant; lot-wise test cert",
         "Bolt/nut MTC","On receipt","H"),
        ("1.5","Welder qualification record (WQR) verification",
         "All welders qualified per IS 7310 / IS 817; records current (< 3 years)",
         "WQR register","Before fabrication start","H"),
    ],
    "2. CUT, DRILL AND FIT-UP INSPECTION":[
        ("2.1","Dimensional accuracy of cut members",
         "Length tolerance +/-2 mm; angle of cut +/-0.5 deg; check with calibrated steel tape",
         "Cutting inspection report","After cutting","W"),
        ("2.2","Hole diameter, pitch, and edge distance",
         "Hole dia = bolt dia + 2 mm clearance; pitch and edge distance per IS 800 Cl. 10.4",
         "Drilling inspection report","After drilling","W"),
        ("2.3","Fit-up inspection before welding",
         "Root gap <= 2 mm; bevel angle per WPS; misalignment <= 2 mm",
         "Fit-up inspection report","Before welding","H"),
        ("2.4","Weld joint geometry check",
         "Throat thickness, leg length per WPS; visual check weld profile",
         "Weld visual report","During / after welding","W"),
        ("2.5","Weld seam NDT - UT / MPI / DPT",
         "All full-penetration butt welds: UT per IS 4260; fillet welds: MPI or DPT 10% sampling",
         "NDT report (third-party)","After welding","H"),
        ("2.6","Distortion and straightness check after welding",
         "Bow <= L/1000; angular distortion <= 1 deg; twist <= 1 mm/m",
         "Dimensional inspection report","After welding","W"),
    ],
    "3. ASSEMBLY AND SUB-ASSEMBLY INSPECTION":[
        ("3.1","Match-marking of bolt holes in mating flanges",
         "All bolt holes align without reaming; max 3 holes may need minor dressing",
         "Assembly fit-up report","During assembly","W"),
        ("3.2","Flange flatness (mating splice flanges)",
         "Flatness <= 1 mm over 1 m; checked with straight edge + feeler gauge",
         "Flange inspection report","Before dispatch","H"),
        ("3.3","Member length and diagonal dimension check",
         "+/-3 mm on member length; diagonal tolerance +/-5 mm across full section face",
         "Dimensional report","Assembly","W"),
        ("3.4","Platform grating assembly and weld inspection",
         "Anti-slip grating welded / clipped; weld quality visual; load test platform if feasible",
         "Platform inspection report","After assembly","W"),
        ("3.5","Ladder / safety cage assembly check",
         "Rung spacing 300 mm +/- 5 mm; cage ring diameter correct; alignment straight",
         "Ladder inspection report","After assembly","W"),
        ("3.6","Trial erection / pre-assembly at factory",
         "All sections erected to full height or representative sub-assembly; photos recorded",
         "Trial erection report + photos","Before dispatch","H"),
    ],
    "4. SURFACE TREATMENT INSPECTION":[
        ("4.1","Surface preparation (blast cleaning) before galvanizing / painting",
         "Sa 2.5 (near white blast) per ISO 8501-1; checked immediately before treatment",
         "Surface prep report","Before treatment","W"),
        ("4.2","Hot-dip galvanizing thickness measurement",
         "Local coating >= 45 micron; mean >= 55 micron (IS 4759 / BS EN ISO 1461); DFT gauge",
         "GI inspection report","After galvanizing","H"),
        ("4.3","Galvanizing visual - runs, bare spots, lumpiness",
         "No bare spots >2 mm2; runs and drips dressed; no cracking of GI coat",
         "Visual GI report","After galvanizing","W"),
        ("4.4","Aviation colour paint DFT (over galvanized surface)",
         "Primer: epoxy zinc phosphate >= 50 micron; Topcoat: polyurethane >= 60 micron DFT",
         "Paint DFT report","After painting","W"),
        ("4.5","Bolt threads protection before dispatch",
         "All threads greased and protected with polyethylene sleeve / thread protector",
         "Packaging inspection","Before dispatch","R"),
    ],
    "5. PRE-DISPATCH INSPECTION (PDI)":[
        ("5.1","Final dimensional check of all sections",
         "Refer drawings; all critical dims confirmed; no missing holes",
         "PDI report","Before loading","H"),
        ("5.2","Quantity verification against BOM / packing list",
         "100% count; part numbers match drawing; no loose items without marking",
         "BOM vs. packing list","Before dispatch","H"),
        ("5.3","Documentation completeness check",
         "Design report, GA drawings, BOM, MTC, GI certs, NDT reports, WQR, test certs",
         "Document checklist","Before dispatch","H"),
        ("5.4","Packing and transport protection",
         "Heavy members on wooden dunnage; fragile items wrapped; bolts in sealed bags labeled",
         "Packing inspection report","Before dispatch","W"),
        ("5.5","Dispatch clearance certificate",
         "Issued jointly by vendor QC and client/TPI representative",
         "Dispatch clearance cert","Before loading","H"),
    ],
}

row = 4
alt = False
for section, items in qc_data.items():
    section_header(ws_qc, row, 2, 8, f"  {section}")
    row += 1
    for item in items:
        ws_qc.row_dimensions[row].height = 50
        bg = C_SKY if alt else C_WHITE
        hold_color = {"H": C_RED_DARK,"W": C_ORANGE,"R": C_GREEN}.get(item[5], "000000")
        for col, val in [(2,item[0]),(3,item[1]),(4,item[2]),(5,item[3]),(6,item[4]),(7,item[5]),(8,"")]:
            cc = ws_qc.cell(row, col)
            cc.value = val
            cc.fill = hex_fill(bg)
            cc.font = body_font(10)
            if col==7:
                cc.font = Font(name="Calibri",size=10,bold=True,color=hold_color)
            cc.alignment = wrap_center("left")
            cc.border = thin_border()
        alt = not alt
        row += 1

row += 1
section_header(ws_qc, row, 2, 8, "  LEGEND FOR INSPECTION HOLD POINTS", C_DARK_GREY)
row += 1
for leg_item, leg_color, leg_desc in [
    ("H - HOLD POINT", C_RED_DARK,
     "Work CANNOT proceed without client/TPI witness and sign-off. Inspector must be present."),
    ("W - WITNESS POINT", C_ORANGE,
     "Client/TPI to be notified 48 hrs in advance. Work may proceed if inspector absent after notice."),
    ("R - REVIEW POINT", C_GREEN,
     "Submit documentation for review. No physical presence required."),
]:
    ws_qc.row_dimensions[row].height = 25
    cc = ws_qc.cell(row, 2)
    cc.value = leg_item
    cc.fill = hex_fill(leg_color)
    cc.font = Font(name="Calibri",size=10,bold=True,color=C_WHITE)
    cc.alignment = wrap_center("left")
    cc.border = thin_border()
    dc = ws_qc.cell(row, 3)
    dc.value = leg_desc
    dc.fill = hex_fill(C_YELLOW_LITE)
    dc.font = body_font(10)
    dc.alignment = wrap_center("left")
    dc.border = thin_border()
    ws_qc.merge_cells(f"C{row}:H{row}")
    row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6 — SURFACE PROTECTION
# ═══════════════════════════════════════════════════════════════════════════
ws_sp = wb.create_sheet("Surface Protection")
ws_sp.sheet_view.showGridLines = False

for col,w in {1:3,2:5,3:38,4:35,5:28,6:22,7:20,8:18,9:3}.items():
    ws_sp.column_dimensions[get_column_letter(col)].width = w

merge_title(ws_sp,"B1:H1","SURFACE PROTECTION - GALVANIZING, COATING AND CORROSION CHECKLIST",
            C_DARK_BLUE,Font(name="Calibri",size=14,bold=True,color=C_WHITE))
merge_title(ws_sp,"B2:H2",
    "Project: Rajasthan WRA Ph-1 | Environment: Arid/Desert | Corrosivity Category: C3 per ISO 9223",
    C_MID_BLUE,Font(name="Calibri",size=10,color=C_WHITE))

r=3
for col,hdr in {2:"Ref",3:"Item / Question",4:"Requirement",
                5:"Vendor Response",6:"Standard",7:"Importance",8:"Compliance"}.items():
    col_header(ws_sp,r,col,hdr)
ws_sp.row_dimensions[r].height=30

sp_data = {
    "1. HOT-DIP GALVANIZING (HDG)":[
        ("1.1","Which standard governs the galvanizing process?",
         "IS 4759:1996 (HDG on structural steel) and/or BS EN ISO 1461:2009",
         "","IS 4759, ISO 1461","Critical"),
        ("1.2","What is the minimum zinc coating thickness (local and mean)?",
         "Local: >= 45 micron; Mean: >= 55 micron (steel > 6 mm thick)",
         "","IS 4759","Critical"),
        ("1.3","What pre-treatment process is used before galvanizing?",
         "Degreasing -> acid pickling (HCl) -> rinsing -> fluxing -> drying",
         "","IS 4759","High"),
        ("1.4","What is the zinc bath composition?",
         "Zinc purity >= 98.5%; temperature 445-455 deg C; dross control maintained",
         "","IS 4759","High"),
        ("1.5","How are internal surfaces of hollow sections galvanized?",
         "Vent holes provided; internal surface also galvanized; confirm hole positions on drawings",
         "","ISO 1461 Annex B","High"),
        ("1.6","How is repair of GI coating carried out (if damage during transport)?",
         "Cold galvanizing compound (zinc-rich paint >95% Zn in DFM) per IS 4759 Cl. 7; min 3 coats",
         "","IS 4759","High"),
        ("1.7","Who performs coating thickness inspection (DFT gauge)?",
         "Vendor QC + independent TPI (BV/DNV/TUV) - calibrated elcometer/DFT gauge",
         "","ISO 2178","Critical"),
        ("1.8","What documentation is provided for galvanizing?",
         "GI certificate per batch: steel weight galvanized, coating thickness readings, date",
         "","IS 4759","Critical"),
    ],
    "2. ADDITIONAL PAINT / COATING SYSTEM (AVIATION MARKING)":[
        ("2.1","What is the paint system specified over galvanized surface?",
         "Sweep blast -> Epoxy zinc phosphate primer 50 micron DFT -> Mid-coat epoxy 40 micron "
         "-> Polyurethane topcoat 60 micron DFT",
         "","ISO 12944 C3","High"),
        ("2.2","What aviation colours are used and what is the gloss level?",
         "International Orange (Red: RAL 3020) and White (RAL 9010); Semi-gloss (Gloss 40-60)",
         "","DGCA CAR Section 4","Critical"),
        ("2.3","What is the minimum expected paint system durability?",
         "Medium durability (M) per ISO 12944 - 5 to 15 years to first maintenance",
         "","ISO 12944-1","High"),
        ("2.4","How is inter-coat adhesion verified?",
         "Cross-cut adhesion test per ISO 2409; pull-off adhesion >= 3 MPa per ISO 4624",
         "","ISO 2409, ISO 4624","High"),
        ("2.5","What touch-up procedure is used for paint damage at site?",
         "Feathering, clean with solvent, apply matching topcoat; provide touch-up kit with mast",
         "","ISO 12944","Medium"),
    ],
    "3. CORROSION PROTECTION OF BOLTS AND HARDWARE":[
        ("3.1","What surface protection is applied to structural bolts and nuts?",
         "Hot-dip galvanized M24+ bolts per IS 1367 Part 13; zinc thickness >= 40 micron",
         "","IS 1367-13","Critical"),
        ("3.2","Are stainless steel bolts used in instrument mounting areas?",
         "SS 316 bolts for all instrument boom mounting, anemometer brackets, sensor mounts",
         "","Client requirement","High"),
        ("3.3","What anti-seize compound is used on bolt threads?",
         "Molybdenum disulphide (MoS2) grease or copper-based anti-seize on all structural bolts",
         "","Client requirement","High"),
        ("3.4","Are spring washers + flat washers used on all bolted connections?",
         "Yes - lock washer or Nordlock washer to prevent self-loosening under dynamic wind load",
         "","IS 800","High"),
        ("3.5","What protection is provided for anchor bolts below ground?",
         "Coal tar epoxy coating >= 300 micron on anchor bolt embedded section; petrolatum tape wrap",
         "","IS 3043","Critical"),
    ],
    "4. CATHODIC / SACRIFICIAL PROTECTION (IF APPLICABLE)":[
        ("4.1","Are sacrificial anodes proposed for below-ground structural steel?",
         "Zinc or magnesium sacrificial anodes bolted to foundation steelwork if soil is aggressive",
         "","IS 13174","Medium"),
        ("4.2","What is the soil corrosivity test result and corrosion risk classification?",
         "Soil resistivity, pH, sulphate, chloride - classify per ISO 9223 / IS recommendations",
         "","ISO 9223","High"),
    ],
}

row = 4
alt = False
for section, items in sp_data.items():
    section_header(ws_sp, row, 2, 8, f"  {section}")
    row += 1
    for item in items:
        ws_sp.row_dimensions[row].height = 50
        bg = C_SKY if alt else C_WHITE
        importance_color = {
            "Critical": C_RED_DARK,"High": C_ORANGE,"Medium": C_AMBER,"Low": C_GREEN
        }.get(item[5],"000000")
        for col, val in [(2,item[0]),(3,item[1]),(4,item[2]),(5,""),(6,item[3]),(7,item[5]),(8,"")]:
            cc = ws_sp.cell(row, col)
            cc.value = val
            cc.fill = hex_fill(bg)
            cc.font = body_font(10)
            if col==7:
                cc.font = Font(name="Calibri",size=10,bold=True,color=importance_color)
            cc.alignment = wrap_center("left")
            cc.border = thin_border()
        alt = not alt
        row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7 — ERECTION AND SAFETY
# ═══════════════════════════════════════════════════════════════════════════
ws_er = wb.create_sheet("Erection and Safety")
ws_er.sheet_view.showGridLines = False

for col,w in {1:3,2:5,3:38,4:35,5:28,6:22,7:20,8:18,9:3}.items():
    ws_er.column_dimensions[get_column_letter(col)].width = w

merge_title(ws_er,"B1:H1","ERECTION, INSTALLATION AND SAFETY CHECKLIST",
            C_DARK_BLUE,Font(name="Calibri",size=14,bold=True,color=C_WHITE))
merge_title(ws_er,"B2:H2",
    "Project: Rajasthan WRA Ph-1 | Erection Method: Mobile Crane | Site: Jaisalmer, Rajasthan",
    C_MID_BLUE,Font(name="Calibri",size=10,color=C_WHITE))

r=3
for col,hdr in {2:"Ref",3:"Erection / Safety Item",4:"Requirement",
                5:"Vendor / Erection Agency Response",6:"Standard",
                7:"Importance",8:"Compliance"}.items():
    col_header(ws_er,r,col,hdr)
ws_er.row_dimensions[r].height=30

erection_data = {
    "1. PRE-ERECTION PLANNING":[
        ("1.1","Is a detailed erection methodology statement (EMS) provided?",
         "Step-by-step erection sequence; crane positions; section lifting scheme; rigging plan",
         "","Client requirement","Critical"),
        ("1.2","What crane capacity is proposed for erection?",
         "Confirm crane capacity sufficient for heaviest lift at maximum radius; crane chart appended",
         "","BS 7121","Critical"),
        ("1.3","Is a site-specific risk assessment / Job Safety Analysis (JSA) prepared?",
         "JSA for each critical task (crane lift, high-work, electrical); reviewed by HSE officer",
         "","IS 3696","Critical"),
        ("1.4","Are ground bearing pressure calculations provided for crane outriggers?",
         "Crane outrigger loads; ground bearing capacity check; matting if required",
         "","Crane manufacturer data","High"),
        ("1.5","What is the erection team qualification and experience?",
         "Erection supervisor: min 5 years lattice tower erection experience; riggers certified",
         "","Client requirement","High"),
        ("1.6","Is a Permit-to-Work (PTW) system in place?",
         "PTW for height work, crane operation, electrical connections; daily toolbox talk",
         "","OHSAS 18001","Critical"),
    ],
    "2. FOUNDATION AND BASE PLATE INSTALLATION":[
        ("2.1","How are anchor bolt positions set out and checked?",
         "Survey instrument (total station / theodolite); check position and level before concrete pour",
         "","IS 1200","Critical"),
        ("2.2","What is the process for checking anchor bolt plumb and projection?",
         "Templates used during concrete pour; projection tolerance +/-5 mm; plumb <= 1 mm/m",
         "","Client requirement","Critical"),
        ("2.3","What grout specification is used under base plates?",
         "Non-shrink cementitious grout (Fosroc / MBT / equivalent); compressive strength >= 60 MPa",
         "","IS 456","Critical"),
        ("2.4","How is the mast base plumb verified after grouting?",
         "Theodolite check in two orthogonal directions; plumb <= 1:1000 at base",
         "","Client requirement","Critical"),
    ],
    "3. SECTION-BY-SECTION ERECTION":[
        ("3.1","What lifting technique is used for each mast section?",
         "Four-point lift with spreader beam; certified slings; SWL marked on all lifting gear",
         "","IS 3832","Critical"),
        ("3.2","How are mast sections aligned and bolted?",
         "Drift pins used for hole alignment; no reaming without approval; torque wrench used",
         "","IS 800 Cl. 10.4","Critical"),
        ("3.3","What is the bolt tightening procedure and torque values?",
         "Snug tight then full tightening (turn-of-nut method or calibrated torque wrench)",
         "","IS 800 Cl. 10.4","Critical"),
        ("3.4","Is a progressive plumb check performed during erection?",
         "Check plumb every 3 sections; correct any deviation before proceeding",
         "","IEC 61400-12-1","Critical"),
        ("3.5","How are temporary guys / bracing used during erection (if any)?",
         "Temporary guys designed for erection wind speed; anchor positions surveyed; released in sequence",
         "","Erection methodology","High"),
        ("3.6","What is the maximum wind speed permitted during erection?",
         "Stop erection if site wind speed > 10 m/s; anemometer on site during erection",
         "","BS 7121 / HSE guideline","Critical"),
    ],
    "4. EARTHING AND LIGHTNING INSTALLATION":[
        ("4.1","How are earth electrodes installed and tested?",
         "Drive rods using portable driver; fill pit with bentonite/coke; test resistance before backfill",
         "","IS 3043","Critical"),
        ("4.2","How are earth conductors connected to mast structure?",
         "Exothermic weld (cadweld) or compression lugs; no painted/GI surface at connection points",
         "","IS 3043","Critical"),
        ("4.3","What test equipment is used for earth resistance measurement?",
         "Fall-of-potential method; calibrated digital earth tester (Kyoritsu / Fluke)",
         "","IS 3043","Critical"),
    ],
    "5. POST-ERECTION CHECKS AND COMMISSIONING":[
        ("5.1","What are the final verticality / plumb acceptance criteria?",
         "Top-of-mast lateral deviation <= H/500 (<= 240 mm for 120 m) in any direction",
         "","IEC 61400-12-1","Critical"),
        ("5.2","Is a final bolt torque check performed after erection?",
         "100% check on all critical bolted connections (splice flanges, base plate); record torque values",
         "","IS 800","Critical"),
        ("5.3","Is a mast visual inspection report produced post-erection?",
         "Inspection checklist: plumb, bolt tightness, GI damage, platform security, ladder continuity",
         "","Client requirement","Critical"),
        ("5.4","Is a handover / as-built documentation package provided?",
         "As-built drawings, erection report, plumb survey, bolt torque records, earth test report",
         "","Client requirement","Critical"),
        ("5.5","Is a site acceptance test (SAT) conducted before instrument commissioning?",
         "Structural SAT: plumb, security, earthing; Instrument SAT: sensor operation, data flow",
         "","Client requirement","Critical"),
    ],
    "6. WORKER SAFETY AND HSE":[
        ("6.1","What fall protection system is used during erection?",
         "Full-body harness (EN 361); double lanyard; fall-arrest devices on safety rail",
         "","IS 3521, EN 363","Critical"),
        ("6.2","Is a rescue plan for height emergencies in place?",
         "Rescue plan documented; rescue kit at mast base; first-aider on site",
         "","IS 3696","Critical"),
        ("6.3","Are all workers covered under ESIC / insurance during erection?",
         "Workmen's compensation policy; ESIC coverage; contractor's all-risk (CAR) insurance",
         "","Factories Act, WCA","Critical"),
        ("6.4","What is the daily toolbox talk and incident reporting procedure?",
         "Daily TBT before work; near-miss / incident report within 24 hours to client HSE",
         "","Client HSE procedure","High"),
        ("6.5","Is a visitor management and site induction procedure in place?",
         "Site induction mandatory before entering erection zone; visitor log maintained",
         "","Client HSE","High"),
    ],
}

row = 4
alt = False
for section, items in erection_data.items():
    section_header(ws_er, row, 2, 8, f"  {section}")
    row += 1
    for item in items:
        ws_er.row_dimensions[row].height = 50
        bg = C_SKY if alt else C_WHITE
        importance_color = {
            "Critical": C_RED_DARK,"High": C_ORANGE,"Medium": C_AMBER,"Low": C_GREEN
        }.get(item[5],"000000")
        for col, val in [(2,item[0]),(3,item[1]),(4,item[2]),(5,""),(6,item[3]),(7,item[5]),(8,"")]:
            cc = ws_er.cell(row, col)
            cc.value = val
            cc.fill = hex_fill(bg)
            cc.font = body_font(10)
            if col==7:
                cc.font = Font(name="Calibri",size=10,bold=True,color=importance_color)
            cc.alignment = wrap_center("left")
            cc.border = thin_border()
        alt = not alt
        row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 8 — INSTRUMENTATION
# ═══════════════════════════════════════════════════════════════════════════
ws_inst = wb.create_sheet("Instrumentation")
ws_inst.sheet_view.showGridLines = False

for col,w in {1:3,2:5,3:38,4:35,5:28,6:22,7:20,8:18,9:3}.items():
    ws_inst.column_dimensions[get_column_letter(col)].width = w

merge_title(ws_inst,"B1:H1","INSTRUMENTATION, DATA LOGGER AND CABLING - SPECIFICATION CHECKLIST",
            C_DARK_BLUE,Font(name="Calibri",size=14,bold=True,color=C_WHITE))
merge_title(ws_inst,"B2:H2",
    "Project: Rajasthan WRA Ph-1 | Standard: IEC 61400-12-1 | Data Logger: NRG Systems / Ammonit / Kintech",
    C_MID_BLUE,Font(name="Calibri",size=10,color=C_WHITE))

r=3
for col,hdr in {2:"Ref",3:"Instrument / Parameter",4:"Specification Required",
                5:"Proposed by Vendor",6:"Standard",7:"Importance",8:"Compliance"}.items():
    col_header(ws_inst,r,col,hdr)
ws_inst.row_dimensions[r].height=30

inst_data = {
    "1. WIND SPEED MEASUREMENT (ANEMOMETERS)":[
        ("1.1","Type of anemometer",
         "Class 1 cup anemometer (IEC 61400-12-1); Thies First Class Advanced / NRG #40C or equivalent",
         "","IEC 61400-12-1","Critical"),
        ("1.2","Measurement heights",
         "Min at: hub height (120m), 2/3 hub (80m), 1/3 hub (40m); redundant sensor at top level",
         "","IEC 61400-12-1 Annex","Critical"),
        ("1.3","Measurement range",
         "0-75 m/s; starting threshold <= 0.5 m/s",
         "","IEC 61400-12-1","Critical"),
        ("1.4","Accuracy / uncertainty",
         "+/- 2% at 12 m/s; calibration certificate traceable to national standard",
         "","IEC 61400-12-1","Critical"),
        ("1.5","Calibration certificate",
         "MEASNET accredited calibration; calibration date < 12 months from installation",
         "","MEASNET","Critical"),
        ("1.6","Boom orientation and mounting",
         "Boom direction: prevailing wind; offset from mast >= 2x mast face width; SS316 mounting",
         "","IEC 61400-12-1 Cl. 6","Critical"),
        ("1.7","Heated anemometers (if icing risk)",
         "Not mandatory for Rajasthan desert; confirm not required based on site temperature history",
         "","Client requirement","Medium"),
    ],
    "2. WIND DIRECTION MEASUREMENT (WIND VANES)":[
        ("2.1","Type of wind vane",
         "Potentiometric or optical encoder wind vane; Thies / NRG 200P or equivalent",
         "","IEC 61400-12-1","Critical"),
        ("2.2","Measurement heights",
         "Min at: top level (120m) and mid level (80m); mounted perpendicular to anemometer boom",
         "","IEC 61400-12-1","Critical"),
        ("2.3","Measurement range and accuracy",
         "0-360 deg (with <=10 deg dead band); accuracy +/-3 deg; resolution <=1 deg",
         "","IEC 61400-12-1","Critical"),
        ("2.4","Dead band orientation",
         "Dead band pointing toward prevailing wind direction to minimise data loss",
         "","IEC 61400-12-1","High"),
    ],
    "3. TEMPERATURE AND PRESSURE SENSORS":[
        ("3.1","Air temperature sensor type and housing",
         "Pt100 or Pt1000 RTD; naturally ventilated radiation shield; Stevenson screen or equivalent",
         "","WMO Guide","High"),
        ("3.2","Temperature measurement heights",
         "At 2 m and at 120 m minimum; aspirated shield preferred at top level",
         "","IEC 61400-12-1","High"),
        ("3.3","Temperature measurement range and accuracy",
         "Range: -20 deg C to +65 deg C; Accuracy: +/-0.3 deg C",
         "","Client requirement","High"),
        ("3.4","Barometric pressure sensor",
         "Capacitive or piezoelectric; range 800-1100 hPa; accuracy +/-0.5 hPa; at base level",
         "","WMO / IEC 61400-12-1","High"),
    ],
    "4. OTHER SENSORS":[
        ("4.1","Relative humidity sensor",
         "Capacitive RH sensor +/-3% accuracy; range 0-100% RH; co-located with temperature",
         "","WMO / Client req.","Medium"),
        ("4.2","Precipitation / rain gauge",
         "Tipping bucket rain gauge (0.2 mm resolution) for data quality flagging",
         "","WMO","Medium"),
        ("4.3","Ice detection sensor (if required)",
         "Not required for Rajasthan; confirm based on T_min records",
         "","Site-specific","Low"),
        ("4.4","Vertical wind component (3D sonic anemometer)",
         "Optional: 3D ultrasonic anemometer at hub height for turbulence intensity measurement",
         "","IEC 61400-12-1 Annex","Medium"),
    ],
    "5. DATA LOGGER AND ACQUISITION SYSTEM":[
        ("5.1","Data logger type and manufacturer",
         "IEC 61400-12-1 compliant; NRG Symphonie / Ammonit Meteo-40 / Kintech KP or equivalent",
         "","IEC 61400-12-1","Critical"),
        ("5.2","Logging intervals",
         "10-minute statistics (mean, std dev, max, min) mandatory; 1-second raw data optional",
         "","IEC 61400-12-1","Critical"),
        ("5.3","Data storage capacity",
         "Onboard memory >= 1 year of 10-min data; removable CF/SD card + USB port",
         "","Client requirement","Critical"),
        ("5.4","Remote data transfer",
         "GPRS / 4G SIM-based data transfer; daily upload to cloud server; SMS alarm capability",
         "","Client requirement","Critical"),
        ("5.5","Data security and encryption",
         "Encrypted data transmission (HTTPS / VPN); tamper-evident logger enclosure",
         "","Client requirement","High"),
        ("5.6","Clock accuracy and time synchronisation",
         "GPS-synchronized clock; accuracy +/- 1 second; auto-sync on upload",
         "","IEC 61400-12-1","Critical"),
        ("5.7","Input channels",
         "Min 16 analog + 8 digital channels; expandable; state channel allocation in bid",
         "","Client requirement","High"),
        ("5.8","Operating temperature range of data logger",
         "-40 deg C to +70 deg C; internally heated enclosure if temperature goes below -10 deg C",
         "","Client requirement","High"),
        ("5.9","Power supply",
         "Solar panel (40 W min) + sealed gel battery (100 Ah min); backup 3-5 days without sun",
         "","Client requirement","Critical"),
        ("5.10","Enclosure IP rating",
         "IP 65 minimum for data logger enclosure; outdoor-rated, lockable, ventilated",
         "","IEC 60529","Critical"),
    ],
    "6. CABLING AND SIGNAL INTEGRITY":[
        ("6.1","What type of signal cable is used from sensors to logger?",
         "Shielded twisted pair (STP) cable; PE outer sheath; UV resistant; rated -40 to +80 deg C",
         "","Client requirement","Critical"),
        ("6.2","How are cables protected on the mast structure?",
         "PVC conduit or SS316 cable tray secured at <= 2 m intervals; UV-stable cable ties",
         "","Client requirement","High"),
        ("6.3","Are cable lengths within sensor specification limits?",
         "Resistance of cable run within logger channel specifications; voltage drop calculated",
         "","Logger datasheet","High"),
        ("6.4","Is shielding earthed at one end only?",
         "Cable shield earthed at logger end only to avoid ground loops",
         "","IEC 61400-12-1","High"),
        ("6.5","Are all cable entries to logger sealed against moisture ingress?",
         "Cable glands (IP68) on all entries; unused entries blanked",
         "","IP 65/68","Critical"),
    ],
    "7. DATA QUALITY AND REPORTING":[
        ("7.1","What data recovery rate is targeted?",
         ">= 95% data recovery (per channel) over the monitoring period",
         "","IEC 61400-12-1","Critical"),
        ("7.2","What data quality flags / filters are applied?",
         "Out-of-range, icing, sensor failure, tower shadow flags per IEC 61400-12-1",
         "","IEC 61400-12-1","Critical"),
        ("7.3","How frequently is the mast inspected and calibrated?",
         "6-monthly site visit: sensor check, re-orientation, calibration verification, logger check",
         "","Client requirement","High"),
        ("7.4","What reporting format is used for wind data?",
         "Standard .txt / .csv (IEC compliant); compatible with WindPRO / WAsP / OpenWind",
         "","IEC 61400-12-1","Critical"),
        ("7.5","Is a mast commissioning report and sensor configuration sheet provided?",
         "Yes - signed commissioning report: serial nos., boom directions, heights, channel allocations",
         "","Client requirement","Critical"),
    ],
}

row = 4
alt = False
for section, items in inst_data.items():
    section_header(ws_inst, row, 2, 8, f"  {section}")
    row += 1
    for item in items:
        ws_inst.row_dimensions[row].height = 45
        bg = C_SKY if alt else C_WHITE
        importance_color = {
            "Critical": C_RED_DARK,"High": C_ORANGE,"Medium": C_AMBER,"Low": C_GREEN
        }.get(item[5],"000000")
        for col, val in [(2,item[0]),(3,item[1]),(4,item[2]),(5,""),(6,item[3]),(7,item[5]),(8,"")]:
            cc = ws_inst.cell(row, col)
            cc.value = val
            cc.fill = hex_fill(bg)
            cc.font = body_font(10)
            if col==7:
                cc.font = Font(name="Calibri",size=10,bold=True,color=importance_color)
            cc.alignment = wrap_center("left")
            cc.border = thin_border()
        alt = not alt
        row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 9 — DOCUMENTATION REQUIREMENTS
# ═══════════════════════════════════════════════════════════════════════════
ws_doc = wb.create_sheet("Documentation")
ws_doc.sheet_view.showGridLines = False

for col,w in {1:3,2:5,3:40,4:30,5:25,6:20,7:18,8:18,9:3}.items():
    ws_doc.column_dimensions[get_column_letter(col)].width = w

merge_title(ws_doc,"B1:H1","DOCUMENTATION AND DELIVERABLE REQUIREMENTS",
            C_DARK_BLUE,Font(name="Calibri",size=14,bold=True,color=C_WHITE))
merge_title(ws_doc,"B2:H2",
    "Project: Rajasthan WRA Ph-1 | All documents in English; drawings in AutoCAD + PDF",
    C_MID_BLUE,Font(name="Calibri",size=10,color=C_WHITE))

r=3
for col,hdr in {2:"Ref",3:"Document / Deliverable",4:"Format / Standard",
                5:"Submission Timing",6:"Copies Required",
                7:"Mandatory?",8:"Received"}.items():
    col_header(ws_doc,r,col,hdr)
ws_doc.row_dimensions[r].height=30

doc_data = {
    "A. DESIGN AND ENGINEERING DOCUMENTS":[
        ("A1","Structural design calculation report (signed by licensed SE)",
         "PDF; stamped by IE(I) member; includes load summary, member checks, deflection, fatigue",
         "With technical bid","2 hard + soft","Mandatory"),
        ("A2","Foundation design calculation report",
         "PDF; includes geotechnical input, bearing capacity, rebar design, anchor bolt design",
         "With technical bid","2 hard + soft","Mandatory"),
        ("A3","General arrangement (GA) drawings - mast elevation, section views, top plan",
         "AutoCAD .dwg + PDF; scale 1:50 or 1:100; all dimensions, material callouts",
         "With technical bid","2 sets + soft","Mandatory"),
        ("A4","Foundation layout and reinforcement drawings",
         "AutoCAD .dwg + PDF; anchor bolt layout, rebar schedule",
         "Within 2 weeks of PO","Soft + 2 hard","Mandatory"),
        ("A5","Erection drawing and sequence diagram",
         "PDF; section assembly sequence, crane positions, rigging details",
         "4 weeks before erection","Soft + 2 hard","Mandatory"),
        ("A6","Earthing and lightning protection scheme drawing",
         "Single line diagram + layout; component schedule",
         "With technical bid","Soft","Mandatory"),
        ("A7","Instrumentation layout drawing (boom locations, cable routing)",
         "AutoCAD + PDF; heights, boom angles, sensor positions, cable routes",
         "With instrumentation bid","Soft + 1 hard","Mandatory"),
        ("A8","Fatigue analysis report",
         "S-N curve method; design life confirmation >= 25 years",
         "With technical bid","Soft","Preferred"),
    ],
    "B. MATERIAL CERTIFICATES":[
        ("B1","Mill Test Certificates (MTC) for all structural steel",
         "Per IS 2062 format; chemical + mechanical; heat number referenced to BOM",
         "Before fabrication / with shipment","Originals","Mandatory"),
        ("B2","MTC for anchor bolts, nuts, washers",
         "IS 1367 grade; lot-wise certificates",
         "Before fabrication","Originals","Mandatory"),
        ("B3","Galvanizing inspection certificate (per batch)",
         "IS 4759 / ISO 1461 format; DFT readings; GI plant name, date, batch",
         "Before dispatch","Originals","Mandatory"),
        ("B4","Paint inspection record (DFT readings)",
         "Each member; min 5 readings; primer + finish coat DFT",
         "Before dispatch","Soft + 1 hard","Mandatory"),
        ("B5","NDT reports (UT / MPI / DPT)",
         "Third-party inspector signed; heat-referenced; pass/fail status per joint",
         "Before dispatch","Originals","Mandatory"),
        ("B6","Sensor / instrument calibration certificates",
         "MEASNET accredited; individual serial-number-based",
         "With instrumentation delivery","Originals","Mandatory"),
        ("B7","Data logger configuration sheet",
         "Channel allocation, scaling factors, serial numbers, firmware version",
         "At commissioning","Soft + 1 hard","Mandatory"),
    ],
    "C. QA/QC AND INSPECTION RECORDS":[
        ("C1","Inspection Test Plan (ITP) - signed by vendor + client/TPI",
         "PDF; all hold/witness/review points; acceptance criteria; responsible parties",
         "2 weeks before fabrication start","Soft + 1 hard","Mandatory"),
        ("C2","Welder Qualification Records (WQR)",
         "IS 7310 format; welder ID, process, position, test date, result",
         "Before fabrication start","Soft","Mandatory"),
        ("C3","Dimensional inspection reports (cutting, drilling, assembly)",
         "Per ITP; vendor QC sign-off; deviations noted with disposition",
         "At each inspection stage","Soft","Mandatory"),
        ("C4","Trial erection / pre-assembly report with photographs",
         "All sections erected; theodolite survey; deviation record; min 20 photographs",
         "Before dispatch","Soft","Mandatory"),
        ("C5","Pre-dispatch inspection (PDI) report signed by client/TPI",
         "Checklist format; hold point sign-off; quantity verification",
         "Before dispatch","Soft + 1 hard","Mandatory"),
        ("C6","Factory Acceptance Test (FAT) report - instrumentation",
         "All sensors and logger tested on bench; results vs. spec; signed",
         "Before instrument dispatch","Soft","Mandatory"),
    ],
    "D. ERECTION AND COMMISSIONING RECORDS":[
        ("D1","Erection completion report",
         "Sequence of erection; dates; crew; crane details; deviations; sign-off",
         "Within 1 week of erection completion","Soft + 1 hard","Mandatory"),
        ("D2","Plumb / verticality survey report (post-erection)",
         "Total station survey; deviation in X and Y at each section height; vs acceptance criteria",
         "Within 3 days of erection","Soft + 1 hard","Mandatory"),
        ("D3","Bolt torque inspection record (post-erection)",
         "Torque values for all critical joints; 100% check at splices",
         "Within 1 week of erection","Soft","Mandatory"),
        ("D4","Earthing system test report",
         "Earth resistance measurement; method (fall of potential); test equipment calibration cert",
         "Before commissioning","Soft + 1 hard","Mandatory"),
        ("D5","Instrument commissioning / SAT report",
         "Sensor operational check; boom direction; heights confirmed; data logger live test",
         "At commissioning","Soft + 1 hard","Mandatory"),
        ("D6","As-built drawings",
         "AutoCAD + PDF; updated from IFC drawings; any site deviations marked",
         "Within 4 weeks of commissioning","Soft + 1 hard","Mandatory"),
        ("D7","Operation and Maintenance (O&M) manual for mast and instruments",
         "Maintenance schedule; inspection checklist; sensor replacement guide; safety instructions",
         "At commissioning","Soft + 1 hard","Mandatory"),
    ],
    "E. COMMERCIAL AND LEGAL DOCUMENTS":[
        ("E1","Performance Bank Guarantee (PBG)",
         "10% of PO value; unconditional; from scheduled bank; validity = PO duration + 6 months",
         "Within 15 days of PO","Original","Mandatory"),
        ("E2","Insurance certificates (CAR, workmen's compensation)",
         "CAR for full PO value; WCA for all workers; copies before site mobilization",
         "Before site mobilization","Copies","Mandatory"),
        ("E3","Warranty letter / certificate",
         "Minimum 3 years for structure; 1 year for instruments; defects liability clause",
         "At delivery","Original","Mandatory"),
        ("E4","DGCA height clearance application support",
         "Vendor to provide mast height, coordinates for DGCA NOC application by client",
         "Before installation","Soft","Mandatory"),
    ],
}

row = 4
alt = False
for section, items in doc_data.items():
    section_header(ws_doc, row, 2, 8, f"  {section}")
    row += 1
    for item in items:
        ws_doc.row_dimensions[row].height = 45
        bg = C_SKY if alt else C_WHITE
        for col, val in [(2,item[0]),(3,item[1]),(4,item[2]),(5,item[3]),(6,item[4]),(7,item[5]),(8,"")]:
            cc = ws_doc.cell(row, col)
            cc.value = val
            cc.fill = hex_fill(bg)
            cc.font = body_font(10)
            if col==7:
                cc.font = Font(name="Calibri",size=10,bold=True,
                               color=C_RED_DARK if val=="Mandatory" else C_GREEN)
            cc.alignment = wrap_center("left")
            cc.border = thin_border()
        alt = not alt
        row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 10 — COMMERCIAL TERMS
# ═══════════════════════════════════════════════════════════════════════════
ws_com = wb.create_sheet("Commercial Terms")
ws_com.sheet_view.showGridLines = False

for col,w in {1:3,2:5,3:38,4:35,5:28,6:22,7:20,8:18,9:3}.items():
    ws_com.column_dimensions[get_column_letter(col)].width = w

merge_title(ws_com,"B1:H1","COMMERCIAL TERMS, WARRANTIES AND CONTRACTUAL REQUIREMENTS",
            C_DARK_BLUE,Font(name="Calibri",size=14,bold=True,color=C_WHITE))
merge_title(ws_com,"B2:H2",
    "Project: Rajasthan WRA Ph-1 | Currency: INR | Incoterms: DDP Site",
    C_MID_BLUE,Font(name="Calibri",size=10,color=C_WHITE))

r=3
for col,hdr in {2:"Ref",3:"Commercial Clause",4:"Client Requirement",
                5:"Vendor Offered Terms",6:"Remarks",7:"Priority",8:"Agreed?"}.items():
    col_header(ws_com,r,col,hdr)
ws_com.row_dimensions[r].height=30

commercial_data = {
    "1. PRICING AND SCOPE OF SUPPLY":[
        ("1.1","What is the scope of supply - structure only or complete turnkey?",
         "Preferred: Turnkey (supply, transport, erection, commissioning, 1st year O&M)",
         "","Confirm included/excluded items","Critical"),
        ("1.2","Is the price firm or subject to escalation?",
         "Firm price for entire contract duration; no escalation for steel / labour",
         "","Client preference","High"),
        ("1.3","What are the payment milestone terms?",
         "10% advance on PO; 30% on drawing approval; 40% on dispatch; 15% on commissioning; 5% on DLP end",
         "","Client standard terms","Critical"),
        ("1.4","What taxes and duties are included in the price?",
         "GST as applicable; HSN code for lattice mast structure to be confirmed; no hidden costs",
         "","GST Act 2017","High"),
        ("1.5","Is a detailed BoQ / BoM attached to the bid?",
         "Yes - itemized BoQ with unit rates for structure, hardware, instruments, civil, erection",
         "","RFQ requirement","Critical"),
    ],
    "2. DELIVERY AND LOGISTICS":[
        ("2.1","What is the guaranteed delivery period?",
         "Structure: 14 weeks from PO; Instruments: 8 weeks from PO",
         "","Client programme","Critical"),
        ("2.2","What is the delivery basis (Incoterms)?",
         "DDP (Delivered Duty Paid) - delivered to site, Jaisalmer, Rajasthan",
         "","Incoterms 2020","Critical"),
        ("2.3","Is special transport / over-dimensional consignment (ODC) arranged?",
         "Vendor responsible for all transport permits (ODC), escort, police clearance",
         "","Client requirement","High"),
        ("2.4","What is the packing standard for transport?",
         "Wooden cradles for long members; polythene wrap on threaded ends; section-wise color coding",
         "","Client requirement","High"),
        ("2.5","What is the breakage / damage risk in transit responsibility?",
         "Marine/transit insurance by vendor; replacement of damaged parts at vendor's cost",
         "","Commercial terms","Critical"),
    ],
    "3. LIQUIDATED DAMAGES (LD)":[
        ("3.1","What is the LD rate for delay in delivery?",
         "0.5% of PO value per week of delay; maximum 5% of PO value",
         "","Client standard","Critical"),
        ("3.2","What is the LD rate for delay in commissioning?",
         "0.5% per week; max 5% of commissioning milestone value",
         "","Client standard","Critical"),
        ("3.3","What is the LD rate for performance (data recovery < 95%)?",
         "Proportional deduction from monthly O&M fee; re-inspection at vendor's cost",
         "","Client requirement","High"),
    ],
    "4. WARRANTY AND DEFECT LIABILITY":[
        ("4.1","What is the warranty period for mast structure?",
         "Minimum 3 years from commissioning; free repair/replacement of structural defects",
         "","Client requirement","Critical"),
        ("4.2","What is the warranty period for instruments and data logger?",
         "Minimum 1 year from commissioning; manufacturer's warranty passthrough",
         "","Client requirement","Critical"),
        ("4.3","What is the Defects Liability Period (DLP)?",
         "12 months from commissioning; 5% retention released at end of DLP",
         "","Client standard","Critical"),
        ("4.4","What is the vendor's on-call response time for critical failures?",
         "Critical (data loss, mast damage): 48-hour response; Non-critical: 7 days",
         "","Client requirement","High"),
        ("4.5","Is a spare parts list and pricing provided?",
         "Yes - mandatory list: sensors, booms, bolts, cable, logger battery, GI repair kit",
         "","Client requirement","High"),
    ],
    "5. INSURANCE AND INDEMNITY":[
        ("5.1","Is contractor's all-risk (CAR) insurance mandatory?",
         "Yes - insured value = PO value; valid from site mobilization to commissioning",
         "","Client requirement","Critical"),
        ("5.2","Is third-party liability insurance required?",
         "Yes - minimum INR 1 Crore per occurrence; vendor to name client as additional insured",
         "","Client requirement","Critical"),
        ("5.3","Is vendor responsible for any damage to public / private property?",
         "Yes - full indemnity clause; vendor liable for third-party claims from their work",
         "","Contract clause","Critical"),
    ],
    "6. INTELLECTUAL PROPERTY AND CONFIDENTIALITY":[
        ("6.1","Are design drawings and reports confidential to this project?",
         "Yes - NDA to be signed; drawings not to be reused for other projects without consent",
         "","NDA / Contract","High"),
        ("6.2","Who retains IP of the mast structural design?",
         "Client retains right to use for this project; vendor retains design IP for other use",
         "","Contract clause","High"),
    ],
    "7. DISPUTE RESOLUTION":[
        ("7.1","What is the governing law and jurisdiction?",
         "Laws of India; jurisdiction - courts of Jaipur / Mumbai (as agreed)",
         "","Contract","High"),
        ("7.2","Is arbitration preferred for disputes?",
         "Arbitration per Arbitration and Conciliation Act 1996; seat: Jaipur; ICADR rules",
         "","Arbitration Act 1996","High"),
    ],
}

row = 4
alt = False
for section, items in commercial_data.items():
    section_header(ws_com, row, 2, 8, f"  {section}")
    row += 1
    for item in items:
        ws_com.row_dimensions[row].height = 45
        bg = C_SKY if alt else C_WHITE
        importance_color = {
            "Critical": C_RED_DARK,"High": C_ORANGE,"Medium": C_AMBER,"Low": C_GREEN
        }.get(item[4],"000000")
        for col, val in [(2,item[0]),(3,item[1]),(4,item[2]),(5,""),(6,item[3]),(7,item[4]),(8,"")]:
            cc = ws_com.cell(row, col)
            cc.value = val
            cc.fill = hex_fill(bg)
            cc.font = body_font(10)
            if col==7:
                cc.font = Font(name="Calibri",size=10,bold=True,color=importance_color)
            cc.alignment = wrap_center("left")
            cc.border = thin_border()
        alt = not alt
        row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 11 — BID EVALUATION MATRIX
# ═══════════════════════════════════════════════════════════════════════════
ws_bev = wb.create_sheet("Bid Evaluation Matrix")
ws_bev.sheet_view.showGridLines = False

for col,w in {1:3,2:5,3:35,4:12,5:14,6:14,7:14,8:14,9:3}.items():
    ws_bev.column_dimensions[get_column_letter(col)].width = w

merge_title(ws_bev,"B1:H1","TECHNICAL BID EVALUATION MATRIX - MET MAST (LATTICE STRUCTURE)",
            C_DARK_BLUE,Font(name="Calibri",size=14,bold=True,color=C_WHITE))
merge_title(ws_bev,"B2:H2",
    "Project: Rajasthan WRA Ph-1 | Evaluation: 70% Technical / 30% Commercial | Min Technical Score: 65%",
    C_MID_BLUE,Font(name="Calibri",size=10,color=C_WHITE))
merge_title(ws_bev,"B3:H3",
    "Scoring: 0 = Not Compliant | 1 = Partially Compliant | 2 = Mostly Compliant | 3 = Fully Compliant",
    C_YELLOW_LITE,Font(name="Calibri",size=9,italic=True,color=C_DARK_GREY))

vendors = ["Vendor A\n(Ahmedabad)", "Vendor B\n(Pune)", "Vendor C\n(Chennai)"]

r=4
col_header(ws_bev,r,2,"Ref")
col_header(ws_bev,r,3,"Evaluation Criterion")
col_header(ws_bev,r,4,"Max\nScore")
for i,v in enumerate(vendors):
    col_header(ws_bev,r,5+i,v)
col_header(ws_bev,r,8,"Remarks")
ws_bev.row_dimensions[r].height=45

eval_sections = {
    "SECTION 1: COMPANY QUALIFICATION (Weight: 15%)":[
        ("1.1","Years of experience in lattice tower / met mast fabrication",5),
        ("1.2","Number of met masts >=80 m supplied in last 5 years",5),
        ("1.3","ISO 9001 and welding quality certifications",5),
        ("1.4","Financial strength (turnover, net worth)",5),
        ("1.5","HSE record and HSE management system",5),
    ],
    "SECTION 2: MANUFACTURING FACILITY (Weight: 15%)":[
        ("2.1","Workshop area and equipment capability",5),
        ("2.2","In-house galvanizing or accredited tie-up",5),
        ("2.3","NDT capability (in-house / third-party)",5),
        ("2.4","Trial erection facility",5),
        ("2.5","Material traceability system",5),
    ],
    "SECTION 3: STRUCTURAL DESIGN (Weight: 25%)":[
        ("3.1","Design wind speed and IS 875 Part 3 compliance",5),
        ("3.2","Completeness of structural calculation report",5),
        ("3.3","Deflection and tilt criteria compliance",5),
        ("3.4","Foundation design adequacy",5),
        ("3.5","Fatigue analysis provided",5),
        ("3.6","Seismic load consideration (IS 1893)",5),
        ("3.7","Base width / stability ratio",5),
        ("3.8","Lightning protection and earthing design",5),
        ("3.9","Drawing quality and completeness",5),
        ("3.10","Erection methodology statement quality",5),
    ],
    "SECTION 4: MATERIAL AND SURFACE PROTECTION (Weight: 15%)":[
        ("4.1","Steel grade and source (primary mill)",5),
        ("4.2","Bolt grade and specification",5),
        ("4.3","Galvanizing specification (IS 4759 / ISO 1461)",5),
        ("4.4","Paint system for aviation marking",5),
        ("4.5","Corrosion protection below ground",5),
    ],
    "SECTION 5: INSTRUMENTATION (Weight: 15%)":[
        ("5.1","Anemometer class and MEASNET calibration",5),
        ("5.2","Data logger compliance (IEC 61400-12-1)",5),
        ("5.3","Communication and remote data transfer",5),
        ("5.4","Solar power supply adequacy",5),
        ("5.5","Data quality management and reporting",5),
    ],
    "SECTION 6: DELIVERY AND COMMERCIAL (Weight: 15%)":[
        ("6.1","Delivery period (structure and instruments)",5),
        ("6.2","Warranty terms (structure 3 yr, instruments 1 yr)",5),
        ("6.3","Price competitiveness (normalised to L1)",5),
        ("6.4","LD and performance guarantees",5),
        ("6.5","Past client references and feedback",5),
    ],
}

row = 5
alt = False
total_max = 0
for section, items in eval_sections.items():
    section_header(ws_bev, row, 2, 8, f"  {section}", bg=C_DARK_GREY)
    row += 1
    for item in items:
        ws_bev.row_dimensions[row].height = 28
        bg = C_SKY if alt else C_WHITE
        total_max += item[2]
        for col, val in [(2,item[0]),(3,item[1]),(4,item[2]),(5,""),(6,""),(7,""),(8,"")]:
            cc = ws_bev.cell(row, col)
            cc.value = val
            cc.fill = hex_fill(bg)
            cc.font = body_font(10, bold=(col==4))
            if col==4:
                cc.font = Font(name="Calibri",size=10,bold=True,color=C_MID_BLUE)
            cc.alignment = wrap_center("center" if col in [2,4,5,6,7] else "left")
            cc.border = thin_border()
        alt = not alt
        row += 1

# Total row
ws_bev.row_dimensions[row].height = 30
ws_bev.merge_cells(f"B{row}:C{row}")
tc = ws_bev.cell(row, 2)
tc.value = "TOTAL MAXIMUM SCORE"
tc.fill = hex_fill(C_DARK_BLUE)
tc.font = Font(name="Calibri",size=11,bold=True,color=C_WHITE)
tc.alignment = wrap_center("center")
tc.border = thick_border()

ws_bev.cell(row,4).value = total_max
ws_bev.cell(row,4).fill = hex_fill(C_DARK_BLUE)
ws_bev.cell(row,4).font = Font(name="Calibri",size=11,bold=True,color=C_WHITE)
ws_bev.cell(row,4).alignment = wrap_center()
ws_bev.cell(row,4).border = thick_border()

for col in [5,6,7,8]:
    ws_bev.cell(row,col).fill = hex_fill(C_MID_BLUE)
    ws_bev.cell(row,col).value = "SUM"
    ws_bev.cell(row,col).font = Font(name="Calibri",size=10,bold=True,color=C_WHITE)
    ws_bev.cell(row,col).alignment = wrap_center()
    ws_bev.cell(row,col).border = thick_border()

row += 2
section_header(ws_bev, row, 2, 8, "  SCORING GUIDE", C_DARK_GREY)
row += 1
score_guide = [
    ("Score 0","Not Compliant - No evidence provided; requirement not met",C_RED_DARK),
    ("Score 1","Partially Compliant - Some evidence but gaps; partially meets requirement",C_ORANGE),
    ("Score 2","Mostly Compliant - Meets most requirements; minor gaps noted",C_AMBER),
    ("Score 3","Fully Compliant - Complete evidence; fully meets or exceeds requirement",C_GREEN),
]
for sg in score_guide:
    ws_bev.row_dimensions[row].height = 22
    cc2 = ws_bev.cell(row, 2)
    cc2.value = sg[0]
    cc2.fill = hex_fill(sg[2])
    cc2.font = Font(name="Calibri",size=10,bold=True,color=C_WHITE)
    cc2.alignment = wrap_center()
    cc2.border = thin_border()
    cc3 = ws_bev.cell(row, 3)
    cc3.value = sg[1]
    cc3.fill = hex_fill(C_YELLOW_LITE)
    cc3.font = body_font(10)
    cc3.alignment = wrap_center("left")
    cc3.border = thin_border()
    ws_bev.merge_cells(f"C{row}:H{row}")
    row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════
output_path = "/home/z/my-project/download/Met_Mast_Procurement_Checklist_India.xlsx"
# For Mac/Linux use:
# output_path = "/Users/YourName/Desktop/Met_Mast_Procurement_Checklist_India.xlsx"
# For Google Colab use:
# output_path = "/content/Met_Mast_Procurement_Checklist_India.xlsx"

wb.save(output_path)
print("File saved successfully!")
print(f"Location: {output_path}")
print(f"Sheets created: {[ws.title for ws in wb.worksheets]}")
print(f"Total sheets: {len(wb.worksheets)}")